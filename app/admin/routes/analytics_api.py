"""
API роуты для системы аналитики
Предоставляет эндпоинты для получения различных метрик и статистики
"""
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Response, Request
from sqlalchemy.orm import Session
import io
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database.connection import get_db
from app.services.analytics_service import AnalyticsService
from app.admin.auth import verify_admin_api


router = APIRouter(tags=["analytics"])


@router.get("/overview")
async def get_overview_analytics(
    request: Request,
    db: Session = Depends(get_db)
):
    """Получить общий обзор аналитики"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    return analytics_service.get_overview_stats()


@router.get("/orders/status")
async def get_orders_by_status(
    request: Request,
    db: Session = Depends(get_db)
):
    """Получить распределение заказов по статусам"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    return analytics_service.get_orders_by_status()


@router.get("/orders/timeline")
async def get_orders_timeline(
    request: Request,
    days: int = 30,
    db: Session = Depends(get_db)
):
    """Получить временную динамику заказов"""
    verify_admin_api(request)
    if days > 365:
        raise HTTPException(status_code=400, detail="Максимальный период - 365 дней")
    
    analytics_service = AnalyticsService(db)
    return analytics_service.get_orders_timeline(days)


@router.get("/revenue/timeline")
async def get_revenue_timeline(
    request: Request,
    days: int = 30,
    db: Session = Depends(get_db)
):
    """Получить временную динамику доходов"""
    verify_admin_api(request)
    if days > 365:
        raise HTTPException(status_code=400, detail="Максимальный период - 365 дней")
    
    analytics_service = AnalyticsService(db)
    return analytics_service.get_revenue_timeline(days)


@router.get("/conversion/funnel")
async def get_conversion_funnel(
    request: Request,
    db: Session = Depends(get_db)
):
    """Получить воронку конверсии заказов"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    return analytics_service.get_conversion_funnel()


@router.get("/work-types")
async def get_work_type_analytics(
    request: Request,
    db: Session = Depends(get_db)
):
    """Получить аналитику по типам работ"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    return analytics_service.get_work_type_analytics()


@router.get("/users")
async def get_user_analytics(
    request: Request,
    db: Session = Depends(get_db)
):
    """Получить аналитику по пользователям"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    return analytics_service.get_user_analytics()


@router.get("/financial")
async def get_financial_analytics(
    request: Request,
    db: Session = Depends(get_db)
):
    """Получить финансовую аналитику"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    return analytics_service.get_financial_analytics()


@router.get("/operational")
async def get_operational_metrics(
    request: Request,
    db: Session = Depends(get_db)
):
    """Получить операционные метрики"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    return analytics_service.get_operational_metrics()


@router.get("/realtime")
async def get_realtime_stats(
    request: Request,
    db: Session = Depends(get_db)
):
    """Получить статистику в реальном времени"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    return analytics_service.get_real_time_stats()


@router.get("/export/csv")
async def export_data_csv(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    data_type: str = "orders",  # orders, payments, all
    db: Session = Depends(get_db)
):
    """Экспорт данных в CSV формате"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    
    # Парсинг дат
    start_dt = None
    end_dt = None
    
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        except:
            raise HTTPException(status_code=400, detail="Неверный формат даты start_date")
    
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        except:
            raise HTTPException(status_code=400, detail="Неверный формат даты end_date")
    
    # Получение данных
    export_data = analytics_service.get_export_data(start_dt, end_dt)
    
    # Создание CSV
    output = io.StringIO()
    
    if data_type == "orders" or data_type == "all":
        # CSV для заказов
        if export_data["orders"]:
            fieldnames = export_data["orders"][0].keys()
            writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            writer.writerows(export_data["orders"])
    
    if data_type == "payments":
        # CSV для платежей
        if export_data["payments"]:
            fieldnames = export_data["payments"][0].keys()
            writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            writer.writerows(export_data["payments"])
    
    # Генерация имени файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"analytics_{data_type}_{timestamp}.csv"
    
    # Возврат файла
    content = output.getvalue()
    output.close()
    
    return Response(
        content=content.encode('utf-8-sig'),  # BOM для корректного отображения в Excel
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/excel")
async def export_data_excel(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Экспорт данных в Excel формате"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    
    # Парсинг дат
    start_dt = None
    end_dt = None
    
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        except:
            raise HTTPException(status_code=400, detail="Неверный формат даты start_date")
    
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        except:
            raise HTTPException(status_code=400, detail="Неверный формат даты end_date")
    
    # Получение данных
    export_data = analytics_service.get_export_data(start_dt, end_dt)
    overview_stats = analytics_service.get_overview_stats()
    financial_stats = analytics_service.get_financial_analytics()
    
    # Создание Excel файла
    wb = Workbook()
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Лист со статистикой
    ws_stats = wb.active
    ws_stats.title = "Статистика"
    
    # Заголовки статистики
    stats_data = [
        ["Метрика", "Значение"],
        ["Всего заказов", overview_stats["total_orders"]],
        ["Всего пользователей", overview_stats["total_users"]],
        ["Общий доход", f"{overview_stats['total_revenue']:.2f} ₽"],
        ["Средний чек", f"{overview_stats['average_order_value']:.2f} ₽"],
        ["Процент завершения", f"{overview_stats['completion_rate']:.2f}%"],
        ["Доход за месяц", f"{financial_stats['monthly_revenue']:.2f} ₽"],
        ["Средний платеж", f"{financial_stats['average_payment']:.2f} ₽"],
    ]
    
    for row_idx, row_data in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Заголовок
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
    
    # Автоширина колонок
    for col in ws_stats.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_stats.column_dimensions[column].width = adjusted_width
    
    # Лист с заказами
    if export_data["orders"]:
        ws_orders = wb.create_sheet("Заказы")
        
        # Заголовки
        headers = list(export_data["orders"][0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws_orders.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Данные
        for row_idx, order in enumerate(export_data["orders"], 2):
            for col_idx, (key, value) in enumerate(order.items(), 1):
                ws_orders.cell(row=row_idx, column=col_idx, value=value)
        
        # Автоширина колонок
        for col in ws_orders.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Максимум 50 символов
            ws_orders.column_dimensions[column].width = adjusted_width
    
    # Лист с платежами
    if export_data["payments"]:
        ws_payments = wb.create_sheet("Платежи")
        
        # Заголовки
        headers = list(export_data["payments"][0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws_payments.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Данные
        for row_idx, payment in enumerate(export_data["payments"], 2):
            for col_idx, (key, value) in enumerate(payment.items(), 1):
                ws_payments.cell(row=row_idx, column=col_idx, value=value)
        
        # Автоширина колонок
        for col in ws_payments.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            ws_payments.column_dimensions[column].width = adjusted_width
    
    # Сохранение в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Генерация имени файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"analytics_report_{timestamp}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/json")
async def export_data_json(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Экспорт данных в JSON формате"""
    verify_admin_api(request)
    analytics_service = AnalyticsService(db)
    
    # Парсинг дат
    start_dt = None
    end_dt = None
    
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        except:
            raise HTTPException(status_code=400, detail="Неверный формат даты start_date")
    
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        except:
            raise HTTPException(status_code=400, detail="Неверный формат даты end_date")
    
    # Получение всех данных
    export_data = analytics_service.get_export_data(start_dt, end_dt)
    overview_stats = analytics_service.get_overview_stats()
    financial_stats = analytics_service.get_financial_analytics()
    operational_stats = analytics_service.get_operational_metrics()
    user_stats = analytics_service.get_user_analytics()
    
    # Объединение всех данных
    full_export = {
        "export_info": {
            "generated_at": datetime.now().isoformat(),
            "start_date": start_date,
            "end_date": end_date,
            "orders_count": len(export_data["orders"]),
            "payments_count": len(export_data["payments"])
        },
        "overview_statistics": overview_stats,
        "financial_analytics": financial_stats,
        "operational_metrics": operational_stats,
        "user_analytics": user_stats,
        "orders": export_data["orders"],
        "payments": export_data["payments"]
    }
    
    # Генерация имени файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"analytics_full_export_{timestamp}.json"
    
    json_content = json.dumps(full_export, ensure_ascii=False, indent=2)
    
    return Response(
        content=json_content.encode('utf-8'),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
